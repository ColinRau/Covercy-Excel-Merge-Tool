import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import difflib
import os
from datetime import datetime, date, timedelta


# Page config & branding
st.set_page_config(
    page_title="Covercy‑Style Excel Merge",
    page_icon="logo.png",
    layout="wide",
)

# Custom CSS for Covercy style
st.markdown(
    """
    <style>
      body { background-color: #ffffff; }
      .stButton>button {
        background-color: #FBBF24;
        color: #111827;
        border-radius: 0.5rem;
        padding: 0.6rem 1.2rem;
      }
      h1, h2, h3 { font-family: 'Helvetica Neue', Arial, sans-serif; }
    </style>
    """,
    unsafe_allow_html=True,
)

# Branded header
this_dir = os.path.dirname(os.path.abspath(__file__))
logo_path = os.path.join(this_dir, "logo.png")
col1, col2 = st.columns([1, 5])
with col1:
    st.image(logo_path, width=160)
with col2:
    st.markdown(
        """
        <h1 style=\"margin:0; color:#111827; font-family:Arial, sans-serif;\">
          Covercy Excel Merge Tool
        </h1>
        """,
        unsafe_allow_html=True,
    )
st.markdown("<hr/>", unsafe_allow_html=True)

# Mode selection
tab = st.sidebar.radio(
    "Select mode:",
    ["Distributions: Complete Import File", "Distributions: Incomplete Import File"]
)

# === Complete flow ===
def run_complete_flow():
    st.title("Distributions: Complete Import File")

    # How-to instructions
    with st.expander("How to"):
        st.markdown(
              """
1. Prepare Customer Source Data Spreadsheet  
    a. If there are multiple tabs within the source spreadsheet, ensure the data you want mapped is in the first tab (optionally, you can also bring that data into its own spreadsheet)  
    b. Ensure Distribution Dates, Investor/Investing Entity Names, and Distribution Amounts are each in their own columns  
    c. (optional) replace non-matching investor names in the source sheet, with the exact Investing Entity Name from Covercy's Import Sheet. This application will try to match non-matching names, and will give you the opportunity to match names manually, but it sometimes helps to do this ahead of time  
    d. In a row above where the data begins, write in headers above the relevant data to help find the data you want mapped (e.g. "Date," "Investing Entity," "Amount")  
    e. Close that spreadsheet (it won't upload if the spreadsheet is opened)  
2. Upload Customer Source Data Spreadsheet where it says "Source Excel File"  
3. Upload Covercy Import Sheet where it says "Target Excel File"  
4. Ignore "Source Data Preview" table for now  
5. Choose the Source Sheet columns you want to map data from  
    a. Where it says "Select Investing Entity column" choose your investing entity header  
    b. Where it says "Select Date column" choose your date header  
    c. Where it says "Select Amount column" choose your amount header  
6. Map Target Entities: Once you have chosen which columns to pull data from, you can see what investing entities from the source sheet will be mapped to the corresponding investing entity in the import file  
    a. Column: Source_Entity shows all Data it sees below your Investing Entity Header in the source file. There may be non-investing entities here if there are non-entity names in that column. Don't worry about these. They will not be mapped unless you choose to map them  
    b. Column: suggestion_1 shows investing entities from the import sheet the app thinks are the best match for the corresponding entity name in the source file. These will not impact the final Data Merge  
    c. Column: Target_Entity shows the investing entities from the import file that will actually be mapped to. These are the entities that matter. If there is a name here, any distribution amounts associated with the relevant name in the Source_Entity column will be pulled into the import file for the Target_Entity Investor  
    d. Within the Target_Entity Column, you may delete names, or write in new names (for instance, if the source file shows John Smith and Jon smith (a typo in this case), you may want to make sure that the same investing entity name is written in the associated row in the Target_Entity Column  
    e. The names in the Target_Entity Column must match what is in Covercy's Import file exactly, so if you write in/copy and paste a name, make sure they match!  
7. Resolve Duplicates: If there is more than one distribution amount associated with the same date for the same investing entity, you may choose here to either pick one amount, or sum them together  
    - Important Note: This application cannot yet distinguish between different types of distributions (e.g. Preferred Return vs Return of Capital) so you have to check!  
8. Click "Finalize and Download Updated Target" - This will not actually download the import file yet  
9. Click "Download Updated Target" - This will download the import file  
10. Check and make sure everything looks good! You should still do your best to ensure the data is mapped accurately.  
11. Upload into Covercy.
            """
        )


    # 1) Upload
    source_file = st.file_uploader("Upload source Excel file", type=["xlsx","xls"], key="src")
    target_file = st.file_uploader("Upload target Excel file", type=["xlsx","xls"], key="tgt")
    if not (source_file and target_file):
        return

    # 2) Read & preview source
    df_source = pd.read_excel(source_file)
    st.subheader("Source Data Preview")
    st.dataframe(df_source.head(5))

    # 3) Column mapping
    cols = df_source.columns.tolist()
    src_ent = st.selectbox("Select Investing Entity column", cols)
    src_dt  = st.selectbox("Select Date column", cols)
    src_amt = st.selectbox("Select Amount column", cols)

    df_source['parsed_date'] = pd.to_datetime(df_source[src_dt], errors='coerce').dt.date
    invalid = df_source['parsed_date'].isna().sum()
    if invalid:
        st.warning(f"{invalid} rows have unparseable dates and will be skipped.")

    # 4) Inspect target sheet layout
    df_raw = pd.read_excel(target_file, header=None)
    ent_col = 2
    ent_label_row = df_raw[df_raw[ent_col]== 'Investing Entity'].index[0]
    gp_row        = df_raw[df_raw[ent_col]== 'GP'].index[0]
    ent_rows      = list(range(ent_label_row+1, gp_row))
    target_entities = [str(df_raw.iat[r,ent_col]).strip() for r in ent_rows]

    date_label_row = ent_label_row - 2
    date_cols = [j for j in range(df_raw.shape[1])
                 if str(df_raw.iat[date_label_row,j]).strip()== 'Last Day']
    date_map = {j: pd.to_datetime(df_raw.iat[date_label_row+1,j], errors='coerce').date()
                for j in date_cols}

    # 5) Entity mapping UI
    unique_src = df_source[src_ent].dropna().astype(str).unique().tolist()
    map_df = pd.DataFrame({'source_entity': unique_src})
    map_df['suggestion_1'] = map_df['source_entity'].apply(
        lambda x: (difflib.get_close_matches(x, target_entities, n=1, cutoff=0.6) or [""])[0]
    )
    map_df['target_entity'] = map_df['suggestion_1']

    st.subheader("Map Source Entities to Target Entities")
    edited = st.data_editor(
        map_df,
        column_config={
            'target_entity': {'editable':True,'type':'dropdown','options':target_entities}
        }, hide_index=True
    )

    # 6) Duplicate resolution
    mapping = dict(zip(edited['source_entity'], edited['target_entity']))
    df_source['mapped_entity'] = df_source[src_ent].map(mapping)

    valid_dates = set(date_map.values())
    dup_src = df_source[
        (df_source['mapped_entity'] != "") &
        (df_source['parsed_date'].isin(valid_dates))
    ]
    dup_groups = dup_src.groupby(
        ['mapped_entity', 'parsed_date']
    )[src_amt].apply(list).reset_index(name='amounts')
    dups = dup_groups[dup_groups['amounts'].apply(len) > 1]
    chosen = {}

    if not dups.empty:
        st.subheader("Resolve Duplicate Amounts")

        # — Sum All button —
        if st.button("Sum All Duplicates"):
            for _, row in dups.iterrows():
                key = f"dup_{row['mapped_entity']}_{row['parsed_date']}"
                st.session_state[key] = 'SUM'

        # — Individual radios —
        for _, row in dups.iterrows():
            ent, dt, amts = row['mapped_entity'], row['parsed_date'], row['amounts']
            key = f"dup_{ent}_{dt}"
            options = [str(a) for a in amts] + ['SUM']
            if key not in st.session_state:
                st.session_state[key] = 'SUM'
            # radio uses session_state[key] as its value
            sel = st.radio(f"Select amount for {ent} on {dt}", options, key=key)
            # read back from session_state so "Sum All" overrides persist
            sel = st.session_state[key]
            chosen[(ent, dt)] = sum(amts) if sel == 'SUM' else float(sel)

    # 7. Finalize and write-back…
    if st.button("Finalize and Download Updated Target"):
        # … your existing write-back logic …

        wb = load_workbook(filename=target_file)
        ws = wb[wb.sheetnames[0]]
        unmatched=[]
        for r,ent in zip(ent_rows, target_entities):
            for col_idx,dist_date in date_map.items():
                if pd.isna(dist_date): continue
                m = df_source[(df_source['mapped_entity']==ent)&(df_source['parsed_date']==dist_date)]
                if not m.empty:
                    amt = chosen.get((ent,dist_date), m[src_amt].iloc[0])
                    ws.cell(row=r+1, column=col_idx).value = amt
                else:
                    unmatched.append((ent,dist_date))
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        st.success("Updated target workbook successfully!")
        st.download_button(
            "Download Updated Target", data=buf,
            file_name="updated_target.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        if unmatched:
            st.warning(f"{len(unmatched)} entries were not matched.")
            st.write(unmatched)

# === Incomplete flow ===
def run_incomplete_flow():
    st.title("Distributions: Incomplete Import File")

    # How-to instructions
    with st.expander("How to"):
        st.markdown(
             """
Note: this section of the app is still in progress. Read the steps carefully  

Important Note: This application cannot yet distinguish between different types of distributions (e.g. Preferred Return vs Return of Capital) so you should probably separate them when you prepare the source sheet  

Steps:  
1. Generate a Covercy Distribution Import file for the relevant asset with a single, custom distribution date. Preferably, a date with no actual distribution BEFORE any real distribution  
2. Prepare Customer Source Data Spreadsheet  
    a. If there are multiple tabs within the source spreadsheet, ensure the data you want mapped is in the first tab (optionally, you can also bring that data into its own spreadsheet)  
    b. Ensure Distribution Dates, Investor/Investing Entity Names, and Distribution Amounts are each in their own columns  
    c. (optional) replace non-matching investor names in the source sheet, with the exact Investing Entity Name from Covercy's Import Sheet. This application will try to match non-matching names, and will give you the opportunity to match names manually, but it sometimes helps to do this ahead of time  
    d. In a row above where the data begins, write in headers above the relevant data to help find the data you want mapped (e.g. "Date," "Investing Entity," "Amount")  
    e. Close that spreadsheet (it won't upload if the spreadsheet is opened)  
3. Upload Customer Source Data Spreadsheet where it says "Source Excel File"  
4. Upload the Single Distribution Date Covercy Import file where it says "incomplete target file"  
5. Choose which Dates to add is a little buggy. Best thing to do is leave all selected, and scroll down to "Or filter by date range"  
6. By default, the whole range of dates will be selected  
7. If you want to select only a custom range, select the dates you want. you will select both dates back to back, so it may feel a little weird. you should also be able to manually edit the dates. This is a little buggy, and will be improved.  
8. Download Populated Template - Now you will have an Import file Complete with all distribution periods  
9. Make sure Distribution Type is what it should be. Right now, this is best for preferred return.  
10. Go to "Distributions: Complete Import File" and use this new import file as your "Target Excel File"  

**IMPORTANT NOTE:** For some reason, Covercy only seems to accept about the first 60% of the distribution periods. Still working on a fix for this. Therefore after you import, see what the last distribution date uploaded was, come back to this application page, upload the original Covercy Import file again, and start from step 7... this time beginning on the date of the next distribution, and ending on the date of the last distribution. You may do this a few times.  

Supplementary Note: It's helpful to pay attention to step 1, and use a date with no actual distribution BEFORE any real distribution when making the initial import file, so that when you re-upload the same file, you don't end up importing duplicate distributions
            """
        )


    # 1) Upload files
    source_file = st.file_uploader("Upload source Excel file", type=["xlsx","xls"], key="inc_src")
    target_file = st.file_uploader("Upload incomplete target file", type=["xlsx","xls"], key="inc_tgt")
    if not (source_file and target_file):
        return

    # 2) Read & parse source dates
    df_src = pd.read_excel(source_file)
    st.subheader("Source Data Preview (Incomplete Flow)")
    st.dataframe(df_src.head(5))

    # Let the user pick which column holds the dates
    inc_date_col = st.selectbox(
        "Select Date column from source",
        options=df_src.columns.tolist(),
        key="inc_date_col"
    )

    # Parse dates using their choice
    df_src['parsed_date'] = pd.to_datetime(
        df_src[inc_date_col], errors='coerce'
    ).dt.date

    invalid = df_src['parsed_date'].isna().sum()
    if invalid:
        st.warning(f"{invalid} rows have unparseable dates and will be skipped.")


    # 3) Load target
    data = target_file.read()
    wb = load_workbook(io.BytesIO(data), data_only=False)
    ws = wb.active

    # 4) Locate entity rows
    colC = [c.value for c in ws['C']]
    ent_label_row = colC.index("Investing Entity")+1
    gp_row        = colC.index("GP")+1
    entity_rows   = list(range(ent_label_row+1, gp_row+1))

    # 5) Copy first block headers
    first_col, width = 6, 7
    hdr1 = [ws.cell(row=1, column=c).value for c in range(first_col, first_col+width)]
    hdr3 = [ws.cell(row=3, column=c).value for c in range(first_col, first_col+width)]
    hdr5 = [ws.cell(row=5, column=c).value for c in range(first_col, first_col+width)]

    # 6) Build new-dates list
    existing = ws.cell(row=4, column=first_col+1).value
    uniq = df_src['parsed_date'].unique()
    new_dates = sorted([d for d in uniq if pd.notna(d) and d != existing])

    # Date selection UI
    st.subheader("Choose which dates to add:")
    date_strs = [d.strftime("%Y-%m-%d") for d in new_dates]
    if 'sel_dates' not in st.session_state:
        st.session_state['sel_dates'] = date_strs.copy()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Select All Dates"):
            st.session_state['sel_dates'] = date_strs.copy()
    with c2:
        if st.button("Clear All Dates"):
            st.session_state['sel_dates'] = []
    selected = st.multiselect(
        "Select dates:",
        options=date_strs,
        default=date_strs,
        key='sel_dates',
        format_func=lambda x: datetime.strptime(x, "%Y-%m-%d").strftime("%b %d, %Y")
    )
    st.markdown("---")
    st.write("**Or filter by date range:**")
    if new_dates:
        start, end = st.date_input("Date range:", [new_dates[0], new_dates[-1]], key='date_range')
        rng = [d for d in new_dates if start <= d <= end]
        dates_to_use = sorted([d for d in rng if d.strftime("%Y-%m-%d") in selected])
    else:
        dates_to_use = []

    # — New: select which distribution type to apply to all blocks —
    dist_options = [
        "Preferred Return",
        "Interest",
        "Profit",
        "Return of Capital",
        "Principal",
        "Promote",
        "Catch Up",
        "Available Cash (Profit)"
    ]
    dist_type = st.selectbox(
        "Select Distribution Type for all new blocks",
        options=dist_options,
        index=0
    )

    # ── Add enough 01-Jan-2040 placeholders so total ≥ len/0.56 ──
    from math import ceil
    N = len(dates_to_use)
    required = ceil(N / 0.56)
    extra = required - N
    if extra > 0:
        placeholder = date(2040, 1, 1)
        dates_to_use.extend([placeholder] * extra)
        st.info(f"Added {extra} placeholder period(s) dated {placeholder.strftime('%d %b %Y')} to meet the 56% rule.")

    # 7) Append blocks
    for idx,last_day in enumerate(dates_to_use):
        base = first_col + width*(idx+1)
        # headers
        for r,vals in zip([1,3,5],[hdr1,hdr3,hdr5]):
            for j,v in enumerate(vals):
                ws.cell(row=r, column=base+j).value = v
        # row2 short month range
        short = f"{last_day.day} {last_day.strftime('%b')} {last_day.year}"
        ws.cell(row=2, column=base   ).value = f"{short} - {short}"
        ws.cell(row=2, column=base+1 ).value = "Custom"
        ws.cell(row=2, column=base+2 ).value = "-"
        ws.cell(row=2, column=base+3 ).value = datetime.now().year
        # row4 full month text
        full = f"{last_day.day} {last_day.strftime('%B')} {last_day.year}"
        ws.cell(row=4, column=base   ).value = full
        ws.cell(row=4, column=base+1 ).value = full
        ws.cell(row=4, column=base+2 ).value = dist_type
        ws.cell(row=4, column=base+3 ).value = "USD"
        # payment dates
        pay_col = base+5
        dt_val  = datetime(last_day.year,last_day.month,last_day.day)
        for r in entity_rows:
            cell_pd = ws.cell(row=r, column=pay_col)
            cell_pd.value = dt_val
            cell_pd.number_format = 'm/d/yyyy'
        # gp formula
        prom = base+2; let = get_column_letter(prom)
        s,e = entity_rows[0],entity_rows[-2]
        ws.cell(row=entity_rows[-1], column=base).value = f"=SUM({let}{s}:{let}{e})"
        # net formulas
        g,t,p,a,n = base,base+1,base+2,base+3,base+4
        for r in entity_rows[:-1]:
            expr = (f"=SUM({get_column_letter(g)}{r},-"
                    f"{get_column_letter(t)}{r},"
                    f"{get_column_letter(a)}{r},-"
                    f"{get_column_letter(p)}{r})")
            ws.cell(row=r, column=n).value = expr
        r_gp = entity_rows[-1]
        expr_gp = (f"=SUM({get_column_letter(g)}{r_gp},-"
                  f"{get_column_letter(t)}{r_gp},"
                  f"{get_column_letter(a)}{r_gp})")
        ws.cell(row=r_gp, column=n).value = expr_gp

    # 8) Download
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    st.download_button("Download Populated Template", data=buf,
                       file_name="populated_incomplete_filtered.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.info("Populated template generated—now mapping entities & filling amounts.")

# Dispatch
if tab == "Distributions: Complete Import File":
    run_complete_flow()
else:
    run_incomplete_flow()
